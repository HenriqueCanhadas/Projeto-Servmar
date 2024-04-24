import openpyxl
from openpyxl.styles import Font

import openpyxl
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
import customtkinter as ctk
import threading
from PIL import Image
from tkinter import font

def main(novo_caminho):

    caminho_resultado = novo_caminho
    caminho_cetesb = 'Tabelas Consulta/Tabelas/Tabela_Cetesb.xlsx'
    caminho_epa = 'Tabelas Consulta/Tabelas/Tabela_EPA.xlsx'
    caminho_listaholandesa = 'Tabelas Consulta/Tabelas/Tabela_ListaHolandesa.xlsx'
    caminho_conama = 'Tabelas Consulta/Tabelas/Tabela_Conama.xlsx'
    caminho_analise = novo_caminho

    # Carregar as planilhas
    wb_resultado = openpyxl.load_workbook(caminho_resultado)
    wb_cetesb = openpyxl.load_workbook(caminho_cetesb)
    wb_epa = openpyxl.load_workbook(caminho_epa)
    wb_listaholandesa = openpyxl.load_workbook(caminho_listaholandesa)
    wb_conama = openpyxl.load_workbook(caminho_conama)

    from openpyxl.styles import Font

    ordem_planilhas = ctk.StringVar(value='Analise')
    valor_primario = None
    
    def abrir_radiobutton_modal():
        def definir_ordem_e_fechar(value):
            nonlocal ordem_planilhas, escolha_feita
            mapeamento_inverso = {'Cetesb': 'c', 'EPA': 'e', 'Lista Holandesa': 'l', 'Conama-420': 'o'}
            valor_escolhido = mapeamento_inverso.get(value, value.lower())
            ordem_planilhas.set(valor_escolhido)
            print("Ordem selecionada:", ordem_planilhas.get())
            nova_janela.destroy()
            while not escolha_feita:
                escolha_feita = abrir_radiobutton_janela()

        # Criar uma janela modal
        nova_janela = ctk.CTkToplevel()
        nova_janela.title("SERVMAR")
        nova_janela.geometry("250x265")
        nova_janela.grab_set()
        nova_janela.resizable(width=False, height=False)

        # Adicionar radiobuttons personalizados
        opcoes = ['Cetesb', 'EPA', 'Lista Holandesa', 'Conama-420']

        label = ctk.CTkLabel(nova_janela, text="Escolha o Primeiro\n Valor  de Referencia")
        label.pack(pady=10)  # Ajuste a distância vertical entre a label e os Radiobuttons

        for opcao in opcoes:
            ctk.CTkRadioButton(nova_janela, text=opcao, variable=ordem_planilhas, value=opcao).pack(anchor='w', pady=5)

        # Botão para confirmar a escolha
        botao_confirmar = ctk.CTkButton(nova_janela, text="Confirmar", command=lambda: definir_ordem_e_fechar(ordem_planilhas.get()))
        botao_confirmar.pack(pady=20)

        nova_janela.wait_window()
        return escolha_feita == True


    def abrir_radiobutton_janela():
        # Criar uma nova janela para os Radiobuttons
        nova_janela_radiobutton = ctk.CTkToplevel()
        nova_janela_radiobutton.geometry("245x255")
        nova_janela_radiobutton.title("SERVMAR")
        nova_janela_radiobutton.resizable(width=False, height=False)

        # Variável para armazenar a escolha do usuário
        escolha_var = tk.StringVar(value=1)
        # Função para obter a escolha do usuário e atribuir à variável valor_Primario
        def obter_escolha():
            nonlocal valor_primario, escolha_feita
            escolha = escolha_var.get()
            
            print("Entrou")
            print(escolha)

            if ordem_planilhas.get() == 'c':
                if escolha == "Solo Agrícola":
                     valor_primario = 1
                elif escolha == "Solo Residencial":
                    valor_primario = 2
                elif escolha == "Solo Industrial":
                    valor_primario = 3
                elif escolha == "Água subterrânea":
                    valor_primario = 4


    
            elif ordem_planilhas.get() == 'e':
                print("DEFE")
                if escolha == "Res Solo":
                    print("A")
                    valor_primario = 1
                elif escolha == "Água subterrânea":
                    print("B")
                    valor_primario = 2
                elif escolha == "Res Ar":
                    valor_primario = 3
                elif escolha == "Solo para GW 1123":
                    valor_primario = 4
                elif escolha == "Ind Solo":
                    valor_primario = 5
                elif escolha == "Ind Air":
                    valor_primario = 6

            elif ordem_planilhas.get() == 'l':
                if escolha == "Solo Agricola":
                     valor_primario = 1
                elif escolha == "Solo Residencial":
                    valor_primario = 2
                elif escolha == "Agua Subterranea":
                    valor_primario = 3

            elif ordem_planilhas.get() == 'o':
                if escolha == "Solo Prevenção":
                     valor_primario = 1
                elif escolha == "Solo Agricola":
                    valor_primario = 2
                elif escolha == "Solo Residencial":
                    valor_primario = 3
                elif escolha == "Solo Industrial":
                    valor_primario = 4
                elif escolha == "Agua Subterranea":
                    valor_primario = 5
            print("Valor Primário selecionado:", valor_primario)

            escolha_feita = True
            nova_janela_radiobutton.destroy()
            return escolha_feita == True
        
        if ordem_planilhas.get() == 'c':
            label = ctk.CTkLabel(nova_janela_radiobutton, text="Selecione a matriz e/ou\n o cenário ambiental da Cetesb")
            label.pack(pady=10)  # Ajuste a distância vertical entre a label e os Radiobuttons
            opcoes = ["Solo Agrícola", "Solo Residencial", "Solo Industrial", "Água subterrânea"]
        elif ordem_planilhas.get() == 'e':
            print("teste")
            label = ctk.CTkLabel(nova_janela_radiobutton, text="Selecione a matriz e/ou\n o cenário ambiental da EPA")
            label.pack(pady=10)  # Ajuste a distância vertical entre a label e os Radiobuttons
            opcoes = ["Res Solo", "Água subterrânea", "Res Ar", "Solo para GW 1123", "Ind Solo", "Ind Air"]
        elif ordem_planilhas.get() == 'l':
            label = ctk.CTkLabel(nova_janela_radiobutton, text="Selecione a matriz e/ou\n o cenário ambiental da Lista Holandesa")
            label.pack(pady=10)  # Ajuste a distância vertical entre a label e os Radiobuttons
            opcoes = ["Solo Agricola", "Solo Residencial", "Agua Subterranea"]
        elif ordem_planilhas.get() == 'o':
            label = ctk.CTkLabel(nova_janela_radiobutton, text="Selecione a matriz e/ou\n o cenário ambiental da Conama-420")
            label.pack(pady=10)  # Ajuste a distância vertical entre a label e os Radiobuttons
            opcoes = ["Solo Prevenção", "Solo Agricola", "Solo Residencial", "Solo Industrial", "Agua Subterranea"]

        for opcao in opcoes:
            print(opcoes)
            ctk.CTkRadioButton(nova_janela_radiobutton, text=opcao, variable=escolha_var, value=opcao).pack(anchor='w')

        # Botão para confirmar a escolha
        botao_confirmar = ctk.CTkButton(nova_janela_radiobutton, text="Confirmar", command=obter_escolha)
        botao_confirmar.pack(pady=10)

        nova_janela_radiobutton.wait_window()
        return escolha_feita == True


    # Loop até que a escolha seja feita
    escolha_feita = False
    while not escolha_feita:
        escolha_feita = abrir_radiobutton_modal()

    # Restante do código pode continuar a partir daqui, usando os valores atualizados de ordem_planilhas e valor_Primario
    print("Ordem Planilhas:", ordem_planilhas.get())
    print("Valor Primário:", valor_primario)

    print("Ordem Planilhas:", ordem_planilhas.get(), type(ordem_planilhas.get()))
    print("Valor Primário:", valor_primario, type(valor_primario))

    # Iterar sobre as abas da planilha Resultado_Final_Organizado
    for sheet_name in wb_resultado.sheetnames:
        sheet_resultado = wb_resultado[sheet_name]

        # Escolher entre wb_cetesb e wb_epa com base na escolha do usuário
        sheet_escolhido = None
        if ordem_planilhas.get() == 'c':
            sheet_escolhido = wb_cetesb['Sheet1']
        elif ordem_planilhas.get() == 'e':
            sheet_escolhido = wb_epa['Sheet1']
        elif ordem_planilhas.get() == 'l':
            sheet_escolhido = wb_listaholandesa['Sheet1']
        elif ordem_planilhas.get() == 'o':
            sheet_escolhido = wb_conama['Sheet1']

        # Variável para armazenar o índice da coluna onde o primeiro dado é inserido
        indice_coluna_padrao = None

        # Variável para rastrear se o último dado foi lido
        ultimo_dado_lido = False

        # Iterar sobre as linhas da coluna 'CAS' em Resultado_Final_Organizado
        for row_resultado in sheet_resultado.iter_rows(min_row=2, max_row=sheet_resultado.max_row, min_col=2,
                                                        max_col=sheet_resultado.max_column):
            # Obter o valor de 'CAS' em Resultado_Final_Organizado
            cas_value = row_resultado[0].value
            unidade = row_resultado[1].value

            # Verificar se o valor de 'CAS' não é nulo
            if cas_value is not None and unidade is not None:
                # Inicializar uma variável para verificar se houve correspondência
                correspondencia_encontrada = False

                # Inicializar uma lista para armazenar todos os valores da linha correspondente
                valores_associados = []

                # Iterar sobre as linhas da coluna 'CAS' em Tabela_Cetesb ou Tabela_EPA
                for row_primairo in sheet_escolhido.iter_rows(min_row=2, max_row=sheet_escolhido.max_row, min_col=2,
                                                          max_col=sheet_escolhido.max_column):
                    # Verificar se encontrou uma correspondência
                    if row_primairo[0].value == cas_value:
                        # Adicionar todos os valores da linha correspondente à lista
                        valores_associados = [cell.value for cell in row_primairo]

                        # Definir que houve uma correspondência
                        correspondencia_encontrada = True

                        # Sair do loop, pois já encontrou uma correspondência
                        break

                # Se houve correspondência, comparar o quinto item da lista com os valores na mesma linha em Resultado_Final_Organizado.xlsx
                if correspondencia_encontrada and len(valores_associados) >= 1:
                    coluna_selecionada = valores_associados[valor_primario]
                    try:
                        if coluna_selecionada != '-' and coluna_selecionada is not None:
                            coluna_selecionada = float(coluna_selecionada)

                            if unidade == 'mg/L':
                                coluna_selecionada /= 1000
                            elif unidade == 'µg/L':
                                pass
                            else:
                                print(unidade)
                        else:
                            coluna_selecionada = None
                            pass
                        
                    except ValueError:
                        # Se a conversão para int ou float falhar, não faz nada
                        pass
                    

                    # Se esta é a primeira iteração, definir a coluna padrão
                    if indice_coluna_padrao is None:
                        indice_coluna_padrao = sheet_resultado.max_column + 1

                    # Comparar e aplicar a formatação desejada
                    for col_index, cell_resultado in enumerate(row_resultado, start=indice_coluna_padrao):
                        valor_na_mesma_linha = cell_resultado.value

                        try:
                            # Verificar se o valor contém "<" e realizar a comparação
                            if valor_na_mesma_linha is not None and coluna_selecionada is not None:
                                if '<' in str(valor_na_mesma_linha):
                                    cell_resultado.font = Font(color="C0C0C0")  # Pintar o texto de cinza
                                elif float(coluna_selecionada) > float(str(valor_na_mesma_linha).replace(',', '.')):
                                    cell_resultado.font = Font(color="000000")  # Pintar o texto de cinza
                                elif float(coluna_selecionada) < float(str(valor_na_mesma_linha).replace(',', '.')):
                                    cell_resultado.font = Font(color="FF0000")  # Pintar o texto de vermelho
                        except ValueError:
                            # Se a conversão falhar, apenas ignore e continue
                            pass

                    # Adicionar o valor da coluna selecionada como uma nova célula na mesma coluna
                    sheet_resultado.cell(row=row_resultado[0].row, column=indice_coluna_padrao, value=coluna_selecionada)

                else:
                    # Se esta é a primeira iteração, definir a coluna padrão
                    if indice_coluna_padrao is None:
                        indice_coluna_padrao = sheet_resultado.max_column + 1
                    # Adicionar o valor da coluna selecionada como uma nova célula na mesma coluna
                    sheet_resultado.cell(row=row_resultado[0].row, column=indice_coluna_padrao, value='n.e')
                        # Se esta é a primeira iteração, definir a coluna padrão

        if indice_coluna_padrao is None:
            indice_coluna_padrao = sheet_resultado.max_column + 1

        # Marcador de último dado lido
        ultimo_dado_lido = True

        # Imprimir "Valor Cetesb" ou "Valor EPA" uma vez acima do primeiro valor após o último dado
        if ultimo_dado_lido:

            if ordem_planilhas.get() == 'c':
                sheet_resultado.cell(row=2, column=indice_coluna_padrao, value="Valor Cetesb")
            elif ordem_planilhas.get() == 'e':
                sheet_resultado.cell(row=2, column=indice_coluna_padrao, value="Valor EPA")
            elif ordem_planilhas.get() == 'l':
                sheet_resultado.cell(row=2, column=indice_coluna_padrao, value="Valor Lista Holandesa")
            elif ordem_planilhas.get() == 'o':
                sheet_resultado.cell(row=2, column=indice_coluna_padrao, value="Valor Conama")

            # Adicionar "n.e" às células em branco após inserir os valores do último dado lido
            for row_resultado in sheet_resultado.iter_rows(min_row=3, max_row=sheet_resultado.max_row,
                                                            min_col=indice_coluna_padrao, max_col=sheet_resultado.max_column):
                for cell_resultado in row_resultado:
                    if cell_resultado.value is None:
                        cell_resultado.value = 'n.e'

    # Salvar as alterações na planilha Resultado_Final_Organizado.xlsx
    wb_resultado.save(caminho_analise)

    # Carregar novamente a planilha Resultado_Final_Organizado após as alterações
    wb_resultado = openpyxl.load_workbook(caminho_analise)

    ordem_planilhas = ctk.StringVar(value='Analise')
    valor_secundario = None
    
    def abrir_radiobutton_modal():
        def definir_ordem_e_fechar(value):
            nonlocal ordem_planilhas, escolha_feita
            mapeamento_inverso = {'Cetesb': 'c', 'EPA': 'e', 'Lista Holandesa': 'l', 'Conama-420': 'o'}
            valor_escolhido = mapeamento_inverso.get(value, value.lower())
            ordem_planilhas.set(valor_escolhido)
            print("Ordem selecionada:", ordem_planilhas.get())
            nova_janela.destroy()
            while not escolha_feita:
                escolha_feita = abrir_radiobutton_janela()

        # Criar uma janela modal
        nova_janela = ctk.CTkToplevel()
        nova_janela.title("SERVMAR")
        nova_janela.geometry("250x265")
        nova_janela.grab_set()
        nova_janela.resizable(width=False, height=False)

        # Adicionar radiobuttons personalizados
        opcoes = ['Cetesb', 'EPA', 'Lista Holandesa', 'Conama-420']

        label = ctk.CTkLabel(nova_janela, text="Escolha o Segundo\n Valor de Referencia")
        label.pack(pady=10)  # Ajuste a distância vertical entre a label e os Radiobuttons

        for opcao in opcoes:
            ctk.CTkRadioButton(nova_janela, text=opcao, variable=ordem_planilhas, value=opcao).pack(anchor='w', pady=5)

        # Botão para confirmar a escolha
        botao_confirmar = ctk.CTkButton(nova_janela, text="Confirmar", command=lambda: definir_ordem_e_fechar(ordem_planilhas.get()))
        botao_confirmar.pack(pady=20)

        nova_janela.wait_window()
        return escolha_feita == True


    def abrir_radiobutton_janela():
        # Criar uma nova janela para os Radiobuttons
        nova_janela_radiobutton = ctk.CTkToplevel()
        nova_janela_radiobutton.geometry("245x255")
        nova_janela_radiobutton.title("SERVMAR")
        nova_janela_radiobutton.resizable(width=False, height=False)

        # Variável para armazenar a escolha do usuário
        escolha_var = tk.StringVar(value=1)
        # Função para obter a escolha do usuário e atribuir à variável valor_secundario
        def obter_escolha():
            nonlocal valor_secundario, escolha_feita
            escolha = escolha_var.get()

            if ordem_planilhas.get() == 'c':
                if escolha == "Solo Agrícola":
                     valor_secundario = 1
                elif escolha == "Solo Residencial":
                    valor_secundario = 2
                elif escolha == "Solo Industrial":
                    valor_secundario = 3
                elif escolha == "Água subterrânea":
                    valor_secundario = 4

            elif ordem_planilhas.get() == 'e':
                if escolha == "Res Solo":
                     valor_secundario = 1
                elif escolha == "Água subterrânea":
                    valor_secundario = 2
                elif escolha == "Res Ar":
                    valor_secundario = 3
                elif escolha == "Solo para GW 1123":
                    valor_secundario = 4
                elif escolha == "Ind Solo":
                    valor_secundario = 5
                elif escolha == "Ind Air":
                    valor_secundario = 6

            elif ordem_planilhas.get() == 'l':
                if escolha == "Solo Agricola":
                     valor_secundario = 1
                elif escolha == "Solo Residencial":
                    valor_secundario = 2
                elif escolha == "Agua Subterranea":
                    valor_secundario = 3

            elif ordem_planilhas.get() == 'o':
                if escolha == "Solo Prevenção":
                     valor_secundario = 1
                elif escolha == "Solo Agricola":
                    valor_secundario = 2
                elif escolha == "Solo Residencial":
                    valor_secundario = 3
                elif escolha == "Solo Industrial":
                    valor_secundario = 4
                elif escolha == "Agua Subterranea":
                    valor_secundario = 5
            print("Valor Secundario selecionado:", valor_secundario)

            escolha_feita = True
            nova_janela_radiobutton.destroy()
            return escolha_feita == True
        
        if ordem_planilhas.get() == 'c':
            label = ctk.CTkLabel(nova_janela_radiobutton, text="Selecione a matriz e/ou\n o cenário ambiental da Cetesb")
            label.pack(pady=10)  # Ajuste a distância vertical entre a label e os Radiobuttons
            opcoes = ["Solo Agrícola", "Solo Residencial", "Solo Industrial", "Água subterrânea"]
        elif ordem_planilhas.get() == 'e':
            label = ctk.CTkLabel(nova_janela_radiobutton, text="Selecione a matriz e/ou\n o cenário ambiental da EPA")
            label.pack(pady=10)  # Ajuste a distância vertical entre a label e os Radiobuttons
            opcoes = ["Res Solo", "Água subterrânea", "Res Ar", "Solo para GW 1123", "Ind Solo", "Ind Air"]
        elif ordem_planilhas.get() == 'l':
            label = ctk.CTkLabel(nova_janela_radiobutton, text="Selecione a matriz e/ou\n o cenário ambiental da Lista Holandesa")
            label.pack(pady=10)  # Ajuste a distância vertical entre a label e os Radiobuttons
            opcoes = ["Solo Agricola", "Solo Residencial", "Agua Subterranea"]
        elif ordem_planilhas.get() == 'o':
            label = ctk.CTkLabel(nova_janela_radiobutton, text="Selecione a matriz e/ou\n o cenário ambiental da Conama-420")
            label.pack(pady=10)  # Ajuste a distância vertical entre a label e os Radiobuttons
            opcoes = ["Solo Prevenção", "Solo Agricola", "Solo Residencial", "Solo Industrial", "Agua Subterranea"]

        for opcao in opcoes:
            ctk.CTkRadioButton(nova_janela_radiobutton, text=opcao, variable=escolha_var, value=opcao).pack(anchor='w')

        # Botão para confirmar a escolha
        botao_confirmar = ctk.CTkButton(nova_janela_radiobutton, text="Confirmar", command=obter_escolha)
        botao_confirmar.pack(pady=10)

        nova_janela_radiobutton.wait_window()
        return escolha_feita == True


    # Loop até que a escolha seja feita
    escolha_feita = False
    while not escolha_feita:
        escolha_feita = abrir_radiobutton_modal()

    # Restante do código pode continuar a partir daqui, usando os valores atualizados de ordem_planilhas e valor_secundario
    print("Ordem Planilhas:", ordem_planilhas.get())
    print("Valor Primário:", valor_secundario)

    print("Ordem Planilhas:", ordem_planilhas.get(), type(ordem_planilhas.get()))
    print("Valor Primário:", valor_secundario, type(valor_secundario))

    # Iterar sobre as abas da planilha Resultado_Final_Organizado após as alterações
    for sheet_name in wb_resultado.sheetnames:
        sheet_resultado = wb_resultado[sheet_name]
        # Segundo loop para processar 'c' se 'e' foi escolhido no primeiro loop, e vice-versa
        if ordem_planilhas.get() == 'c':
            # Configurar 'c' para o segundo loop
            sheet_escolhido_secundario = wb_cetesb['Sheet1']
        elif ordem_planilhas.get() == 'e':
            # Configurar 'e' para o segundo loop
            sheet_escolhido_secundario = wb_epa['Sheet1']
        elif ordem_planilhas.get() == 'l':
            # Configurar 'e' para o segundo loop
            sheet_escolhido_secundario = wb_listaholandesa['Sheet1']
        elif ordem_planilhas.get() == 'o':
            # Configurar 'e' para o segundo loop
            sheet_escolhido_secundario = wb_conama['Sheet1']

        # Encontrar a coluna "Valor Cetesb" ou "Valor EPA"
        indice_coluna_primaria = None
        for col_index, col in enumerate(sheet_resultado.iter_cols(min_row=2, max_row=2), start=1):
            # Verificar se o valor atual da célula contém "Valor Cetesb" ou "Valor EPA"
            if "Valor Cetesb" in str(col[0].value) or "Valor EPA" in str(col[0].value) or "Valor Lista Holandesa" in str(col[0].value) or "Valor Conama" in str(col[0].value):
                indice_coluna_primaria = col_index
                break

        # Verificar se a coluna "Valor Cetesb" foi encontrada
        # Verificar se a coluna "Valor Cetesb" foi encontrada
        if indice_coluna_primaria is not None:
            # Adicionar uma nova coluna para "Valor EPA" ou "Valor Cetesb" (invertido)
            indice_coluna_secundaria = indice_coluna_primaria + 1

            if ordem_planilhas.get() == 'c':
                titulo_coluna_secundaria = "Valor Cetesb"
            elif ordem_planilhas.get() == 'e':
                titulo_coluna_secundaria = "Valor EPA"
            elif ordem_planilhas.get() == 'l':
                titulo_coluna_secundaria = "Valor Lista Holandesa"
            elif ordem_planilhas.get() == 'o':
                titulo_coluna_secundaria = "Valor Conama"

            sheet_resultado.cell(row=2, column=indice_coluna_secundaria, value=titulo_coluna_secundaria)

            # Iterar sobre as linhas a partir da terceira linha (índice 3)
            for row in range(3, sheet_resultado.max_row + 1):
                valor_primario = sheet_resultado.cell(row=row, column=indice_coluna_primaria).value

                # Comparar e inserir valores na coluna "Valor Secundario"
                if valor_primario in ["n.e"]:

                    cas_value = sheet_resultado.cell(row=row, column=2).value  # Valor da coluna 'CAS'
                    unidade = sheet_resultado.cell(row=row, column=3).value

                    # Verificar se o valor de 'CAS' não é nulo
                    if cas_value is not None and unidade is not None:
                        valores_associados = []  # Inicializar uma lista para armazenar todos os valores da linha correspondente

                        # Iterar sobre as linhas da coluna 'CAS' na planilha secundária (wb_cetesb ou wb_epa)
                        for row_secundaria in sheet_escolhido_secundario.iter_rows(min_row=2, max_row=sheet_escolhido_secundario.max_row, min_col=2, max_col=sheet_escolhido_secundario.max_column):
                            if row_secundaria[0].value == cas_value:
                                valores_associados = [cell.value for cell in row_secundaria]
                                break

                        # Se houve correspondência, comparar o quinto item da lista com os valores na mesma linha em Resultado_Final_Organizado.xlsx
                        # Comparar e inserir valores na coluna "Valor Secundario"
                        if len(valores_associados) >= 1:
                            coluna_selecionada = valores_associados[valor_secundario]

                            try:
                                if coluna_selecionada != '-' and coluna_selecionada is not None:
                                    coluna_selecionada = float(coluna_selecionada)

                                    if unidade == 'mg/L':
                                        coluna_selecionada /= 1000
                                    elif unidade == 'µg/L':
                                        pass
                                    else:
                                        print(unidade)
                                else:
                                    coluna_selecionada = None
                                    pass
                                
                            except ValueError:
                                # Se a conversão para int ou float falhar, não faz nada
                                pass

                            # Se esta é a primeira iteração, definir a coluna padrão
                            if indice_coluna_padrao is None:
                                indice_coluna_padrao = sheet_resultado.max_column + 1

                            # Comparar e aplicar a formatação desejada
                            for col_index in range(2, col_index + 1):
                                cell_resultado = sheet_resultado.cell(row=row, column=col_index)
                                valor_na_mesma_linha = cell_resultado.value

                                try:
                                    # Verificar se a célula não é nula e se o valor contém "<" e realizar a comparação
                                    if valor_na_mesma_linha is not None and coluna_selecionada is not None:
                                        if '<' in str(valor_na_mesma_linha):
                                            cell_resultado.font = Font(color="C0C0C0")  # Pintar o texto de cinza
                                        elif float(coluna_selecionada) > float(str(valor_na_mesma_linha).replace(',', '.')):
                                            cell_resultado.font = Font(color="000000")  # Pintar o texto de cinza
                                        elif float(coluna_selecionada) < float(str(valor_na_mesma_linha).replace(',', '.')):
                                            cell_resultado.font = Font(color="FF0000")  # Pintar o texto de vermelho
                                except ValueError:
                                    # Se a conversão falhar, apenas ignore e continue
                                    pass

                            # Adicionar o valor da coluna selecionada como uma nova célula na mesma coluna
                            sheet_resultado.cell(row=row, column=indice_coluna_secundaria, value=coluna_selecionada)

                            # Marcador de último dado lido
                            ultimo_dado_lido = True
                else:
                    sheet_resultado.cell(row=row, column=indice_coluna_secundaria, value="n.a")

        # Imprimir "Valor Cetesb" ou "Valor EPA" uma vez acima do primeiro valor após o último dado
        if ultimo_dado_lido:
            # Adicionar "n.e" às células em branco após inserir os valores do último dado lido
            for row_resultado in sheet_resultado.iter_rows(min_row=3, max_row=sheet_resultado.max_row,
                                                            min_col=indice_coluna_padrao, max_col=sheet_resultado.max_column):
                for cell_resultado in row_resultado:
                    if cell_resultado.value is None:
                        cell_resultado.value = 'n.e'

    # Salvar as alterações na planilha Resultado_Final_Organizado_Formatado_cetesb_epa.xlsx
    wb_resultado.save(caminho_analise)

    # Carregar novamente a planilha Resultado_Final_Organizado após as alterações
    wb_resultado = openpyxl.load_workbook(caminho_analise)

    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side

    # Carregar novamente a planilha Resultado_Final_Organizado após as alterações
    wb_resultado = openpyxl.load_workbook(caminho_analise)

    # Iterar sobre as abas da planilha Resultado_Final_Organizado após as alterações
    for sheet_name in wb_resultado.sheetnames:
        sheet_resultado = wb_resultado[sheet_name]

        # Inserir uma nova coluna vazia na coluna D
        sheet_resultado.insert_cols(idx=4)
        # Inserir uma nova coluna vazia na coluna E
        sheet_resultado.insert_cols(idx=5)

        # Inicializar variáveis para armazenar as colunas de "Valor Cetesb" e "Valor EPA"
        coluna_cetesb = None
        coluna_epa = None
        coluna_listaholandesa = None
        coluna_conama = None

        # Iterar sobre as células da segunda linha
        for col_idx in range(1, sheet_resultado.max_column + 1):
            valor = sheet_resultado.cell(row=2, column=col_idx).value

            if valor == "Valor Cetesb":
                coluna_cetesb = col_idx
            elif valor == "Valor EPA":
                coluna_epa = col_idx
            elif valor == "Valor Lista Holandesa":
                coluna_listaholandesa = col_idx
            elif valor == "Valor Conama":
                coluna_conama = col_idx

        # Suponha que as variáveis coluna_cetesb, coluna_epa e coluna_listaholandesa já foram definidas anteriormente

        # Encontrar o menor valor entre as variáveis
        menor_valor_entre_colunas = min(
            (coluna_cetesb, 'coluna_cetesb'),
            (coluna_epa, 'coluna_epa'),
            (coluna_listaholandesa, 'coluna_listaholandesa'),
            (coluna_conama, 'coluna_conama'),
            key=lambda x: x[0] if isinstance(x[0], int) else float('inf')
        )

        # Desempacotar o resultado para obter o valor e a variável correspondente
        menor_valor, variavel_menor_valor = menor_valor_entre_colunas

        # Encontrar o maior valor entre as variáveis
        maior_valor_entre_colunas = max(
            (coluna_cetesb, 'coluna_cetesb'),
            (coluna_epa, 'coluna_epa'),
            (coluna_listaholandesa, 'coluna_listaholandesa'),
            (coluna_conama, 'coluna_conama'),
            key=lambda x: x[0] if isinstance(x[0], int) else float('-inf')
        )

        # Desempacotar o resultado para obter o valor e a variável correspondente
        maior_valor, variavel_maior_valor = maior_valor_entre_colunas

        # Copiar os valores para as colunas D e E
        for row_idx in range(2, sheet_resultado.max_row + 1):
                    maior_valor_entre_colunas= sheet_resultado.cell(row=row_idx, column=maior_valor).value
                    menor_valor_entre_colunas = sheet_resultado.cell(row=row_idx, column=menor_valor).value

                    sheet_resultado.cell(row=row_idx, column=4, value=menor_valor_entre_colunas)
                    sheet_resultado.cell(row=row_idx, column=5, value=maior_valor_entre_colunas)

        for row in sheet_resultado.iter_rows(min_row=1, max_row=sheet_resultado.max_row, min_col=menor_valor, max_col=maior_valor):
            for cell in row:
                cell.value = None

        # Inserir o valor "Valor Orientador" na célula D1
        sheet_resultado.cell(row=1, column=4, value="Valor Orientador")

        # Redimensionar automaticamente largura e altura das células
        for col in sheet_resultado.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 1)
            sheet_resultado.column_dimensions[column].width = adjusted_width

        # Alinhar o texto ao centro das células mescladas e aplicar negrito
        for row in sheet_resultado.iter_rows(min_row=1, max_row=sheet_resultado.max_row, min_col=1, max_col=sheet_resultado.max_column - 2):
            for cell in row:

                cell.alignment = Alignment(horizontal='center', vertical='center')
                # Aplicar negrito aos valores "Valor EPA" e "Valor Cetesb"
                if cell.value in ["Valor EPA", "Valor Cetesb","Valor Lista Holandesa","Valor Conama","Valor Orientador"]:
                    cell.font = Font(bold=True)

                # Adicionar todas as bordas com estilo "todas as bordas"
                cell.border = Border(
                    left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000')
                )

        # Mesclar a célula D1 até a célula E1
        sheet_resultado.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)

        # Mesclar células A1 até A2
        sheet_resultado.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        # Mesclar células B1 até B2
        sheet_resultado.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        # Mesclar células C1 até C3
        sheet_resultado.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

        # Verificação e coloração do texto das células a partir da coluna G e linha 3
        for row_idx in range(3, sheet_resultado.max_row + 1):
            for col_idx in range(6, sheet_resultado.max_column + 1):
                cell_value = sheet_resultado.cell(row=row_idx, column=col_idx).value

                # Verifica se o valor na célula é "<" e altera a cor do texto para preto
                if isinstance(cell_value, str) and "<" in cell_value:
                    sheet_resultado.cell(row=row_idx, column=col_idx).font = Font(color="C0C0C0")

    # Salvar as alterações no arquivo
    wb_resultado.save(caminho_analise)

if __name__ == "__main__":
    main()